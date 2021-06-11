#!/usr/local/python3.7.6/bin/python3
# 简单操作
#!/usr/bin/python3
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws['A1'] = "name"
ws['B1'] = "age"
ws['C1'] = "high"
wb.save("results/sample.xlsx")

# xlsx支持导入图片，安装pillow库
pip3 install pillow

# 1.创建一个工作簿
from openpyxl import Workbook
# 工作簿实例
wb = Workbook()

ws = wb.active
# openpyxl.workbook.Workbook.active()

ws1 = wb.create_sheet()
ws2 = wb.create_sheet(0)

ws.title = "New Title"

ws.sheet_properties.tabColor = "1072BA" # 设置背景色